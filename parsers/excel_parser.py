from openpyxl import load_workbook
from pathlib import Path

from constans import CSV_COLUMNS
from utils.additional_info import build_additional_info


EXCEL_ADDITIONAL_FIELDS = {
    "Company": "COMPANY",
    "Department": "DEPARTMENT",
    "Position": "POSITION",
}


def clean_value(value):
    if value is None:
        return ""

    return str(value).strip()


def convert_excel_row_to_csv_row(excel_row):
    result_row = dict.fromkeys(CSV_COLUMNS, "")
        
    first_name = clean_value(excel_row.get("First Name"))
    last_name = clean_value(excel_row.get("Last Name"))
        
    result_row["name"] = first_name
    result_row["user_fulname"] = f"{first_name} {last_name}".strip()
    result_row["address"] = clean_value(excel_row.get("Address"))
    result_row["zip"] = clean_value(excel_row.get("Zip"))
    result_row["tel"] = clean_value(excel_row.get("Mobile number"))
    result_row["type"] = "excel"

    result_row["user_additional_info"] = build_additional_info(
        excel_row,
        EXCEL_ADDITIONAL_FIELDS,
    )

    return result_row


def parse_excel(file_path):
    file_path = Path(file_path)
    
    if not file_path.exists():
        raise FileNotFoundError(f'Excel file not found: {file_path}')
    
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)

    try:
        headers = next(rows)
    except StopIteration:
        return []
    
    parsed_rows = []
    
    for row in rows:
        excel_row = dict(zip(headers, row))
        
        result_row = convert_excel_row_to_csv_row(excel_row)
        
        parsed_rows.append(result_row)

    return parsed_rows
